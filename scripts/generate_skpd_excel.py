"""
Generator Data Penelitian Nasabah SKPD 2022
Menghasilkan data nasabah pembiayaan SKPD lengkap dengan scoring dan kriteria keputusan
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import random
from datetime import date, timedelta
import math

# ─── Referensi Data Master ────────────────────────────────────────────────────

INSTANSI = [
    {"id": 1,  "nama": "BAPPENDA",                     "rating": 4, "bobot": 0.05},
    {"id": 2,  "nama": "BPKSDM",                       "rating": 4, "bobot": 0.05},
    {"id": 3,  "nama": "BPBD",                         "rating": 3, "bobot": 0.05},
    {"id": 4,  "nama": "BPKAD",                        "rating": 3, "bobot": 0.05},
    {"id": 5,  "nama": "DPMPTSP",                      "rating": 3, "bobot": 0.05},
    {"id": 6,  "nama": "DAMKAR",                       "rating": 4, "bobot": 0.05},
    {"id": 7,  "nama": "DINAS ARSIP DAN PERPUSDA",     "rating": 3, "bobot": 0.05},
    {"id": 8,  "nama": "DISKOPUKM",                    "rating": 2, "bobot": 0.05},
    {"id": 9,  "nama": "DP3AP2KB",                     "rating": 4, "bobot": 0.05},
    {"id": 10, "nama": "DPMD",                         "rating": 3, "bobot": 0.05},
    {"id": 11, "nama": "DINAS SOSIAL",                 "rating": 2, "bobot": 0.05},
    {"id": 12, "nama": "DISBUDPAR",                    "rating": 3, "bobot": 0.05},
    {"id": 13, "nama": "DISDUKCAPIL",                  "rating": 3, "bobot": 0.05},
    {"id": 14, "nama": "DISHUB",                       "rating": 4, "bobot": 0.05},
    {"id": 15, "nama": "DISHUB PROVINSI",              "rating": 1, "bobot": 0.05},
    {"id": 16, "nama": "DISKOMINFO",                   "rating": 3, "bobot": 0.05},
    {"id": 17, "nama": "DISNAKER",                     "rating": 3, "bobot": 0.05},
    {"id": 18, "nama": "DISDAGIN",                     "rating": 3, "bobot": 0.05},
    {"id": 19, "nama": "DISPORA",                      "rating": 1, "bobot": 0.05},
    {"id": 20, "nama": "DKDP",                         "rating": 2, "bobot": 0.05},
    {"id": 21, "nama": "DPKPP",                        "rating": 2, "bobot": 0.05},
    {"id": 22, "nama": "DPRD",                         "rating": 2, "bobot": 0.05},
    {"id": 23, "nama": "INSPEKTORAT DAERAH",           "rating": 4, "bobot": 0.05},
    {"id": 24, "nama": "KECAMATAN BABAKAN MADANG",     "rating": 3, "bobot": 0.05},
    {"id": 25, "nama": "KECAMATAN BOJONG GEDE",        "rating": 3, "bobot": 0.05},
    {"id": 26, "nama": "KECAMATAN CARINGIN",           "rating": 2, "bobot": 0.05},
    {"id": 27, "nama": "KECAMATAN CARIU",              "rating": 1, "bobot": 0.05},
    {"id": 28, "nama": "KECAMATAN CIAMPEA",            "rating": 2, "bobot": 0.05},
    {"id": 29, "nama": "KECAMATAN CIAWI",              "rating": 1, "bobot": 0.05},
    {"id": 30, "nama": "KECAMATAN CIBINONG",           "rating": 3, "bobot": 0.05},
    {"id": 31, "nama": "KECAMATAN CIBUNGBULANG",       "rating": 2, "bobot": 0.05},
    {"id": 32, "nama": "KECAMATAN CIGOMBONG",          "rating": 2, "bobot": 0.05},
    {"id": 33, "nama": "KECAMATAN CIOMAS",             "rating": 2, "bobot": 0.05},
    {"id": 34, "nama": "KECAMATAN CISARUA",            "rating": 1, "bobot": 0.05},
    {"id": 35, "nama": "KECAMATAN CISEENG",            "rating": 1, "bobot": 0.05},
    {"id": 36, "nama": "KECAMATAN JONGGOL",            "rating": 1, "bobot": 0.05},
    {"id": 37, "nama": "KECAMATAN KEMANG",             "rating": 1, "bobot": 0.05},
    {"id": 38, "nama": "KECAMATAN LEUWILIANG",         "rating": 3, "bobot": 0.05},
    {"id": 39, "nama": "KECAMATAN MEGAMENDUNG",        "rating": 2, "bobot": 0.05},
    {"id": 40, "nama": "KECAMATAN NANGGUNG",           "rating": 2, "bobot": 0.05},
    {"id": 41, "nama": "KECAMATAN RANCABANGUR",        "rating": 2, "bobot": 0.05},
    {"id": 42, "nama": "KECAMATAN SUKARAJA",           "rating": 1, "bobot": 0.05},
    {"id": 43, "nama": "KECAMATAN TAJURHALANG",        "rating": 2, "bobot": 0.05},
    {"id": 44, "nama": "KECAMATAN TAMAN SARI",         "rating": 1, "bobot": 0.05},
    {"id": 45, "nama": "KECAMATAN TANJUNG SARI",       "rating": 1, "bobot": 0.05},
    {"id": 46, "nama": "KECAMATAN TENJO",              "rating": 1, "bobot": 0.05},
    {"id": 47, "nama": "KECAMATAN CITEUREUP",          "rating": 2, "bobot": 0.05},
    {"id": 48, "nama": "BAKESBANGPOL",                 "rating": 2, "bobot": 0.05},
    {"id": 49, "nama": "SETDA",                        "rating": 4, "bobot": 0.05},
    {"id": 50, "nama": "DINAS PENDIDIKAN",             "rating": 4, "bobot": 0.05},
    {"id": 51, "nama": "DINAS KESEHATAN",              "rating": 4, "bobot": 0.05},
    {"id": 52, "nama": "POL PP",                       "rating": 1, "bobot": 0.05},
    {"id": 53, "nama": "DINAS LINGKUNGAN HIDUP",       "rating": 3, "bobot": 0.05},
    {"id": 54, "nama": "KPU",                          "rating": 2, "bobot": 0.05},
]

BENDAHARA = [
    {"id": 1, "nama": "Tidak Baik dan Tidak Tertib", "rating": 1, "bobot": 0.30},
    {"id": 2, "nama": "Baik dan Tidak Tertib",        "rating": 2, "bobot": 0.30},
    {"id": 3, "nama": "Tidak Baik dan Tertib",        "rating": 3, "bobot": 0.30},
    {"id": 4, "nama": "Baik dan Tertib",              "rating": 4, "bobot": 0.30},
]

GOLONGAN = [
    {"id": 1,  "nama": "Golongan I/a"},  {"id": 2,  "nama": "Golongan I/b"},
    {"id": 3,  "nama": "Golongan I/c"},  {"id": 4,  "nama": "Golongan I/d"},
    {"id": 5,  "nama": "Golongan II/a"}, {"id": 6,  "nama": "Golongan II/b"},
    {"id": 7,  "nama": "Golongan II/c"}, {"id": 8,  "nama": "Golongan II/d"},
    {"id": 9,  "nama": "Golongan III/a"},{"id": 10, "nama": "Golongan III/b"},
    {"id": 11, "nama": "Golongan III/c"},{"id": 12, "nama": "Golongan III/d"},
    {"id": 13, "nama": "Golongan IV/a"}, {"id": 14, "nama": "Golongan IV/b"},
    {"id": 15, "nama": "Golongan IV/c"}, {"id": 16, "nama": "Golongan IV/d"},
    {"id": 17, "nama": "Golongan IV/e"}, {"id": 18, "nama": "Non PNS (Honorer)"},
]

GAJI_BY_GOLONGAN = {
    1: 1560800,  2: 1704500,  3: 1776600,  4: 1851900,
    5: 2022200,  6: 2208400,  7: 2301800,  8: 2399200,
    9: 2579400,  10: 2688500, 11: 2802300, 12: 2920800,
    13: 3042100, 14: 3173100, 15: 3307300, 16: 3447200,
    17: 3593100, 18: 2500000,
}

TPP_BY_GOLONGAN = {
    1: 1200000,  2: 1300000,  3: 1400000,  4: 1500000,
    5: 1800000,  6: 2000000,  7: 2200000,  8: 2400000,
    9: 2800000,  10: 3200000, 11: 3600000, 12: 4000000,
    13: 5000000, 14: 6000000, 15: 7000000, 16: 8000000,
    17: 9000000, 18: 0,
}

STATUS_PERKAWINAN = [
    {"id": 1, "nama": "Belum Menikah",  "biaya": 0},
    {"id": 2, "nama": "Menikah",        "biaya": 500000},
    {"id": 3, "nama": "Duda/Janda",     "biaya": 0},
]

TANGGUNGAN = [
    {"id": 1, "banyak": "Tidak Ada",  "biaya": 0},
    {"id": 2, "banyak": "1 Orang",    "biaya": 400000},
    {"id": 3, "banyak": "2 Orang",    "biaya": 800000},
    {"id": 4, "banyak": "3 Orang",    "biaya": 1200000},
    {"id": 5, "banyak": "> 3 Orang",  "biaya": 1600000},
]

JENIS_NASABAH = [
    {"id": 1, "nama": "Nasabah Baru",              "rating": 2, "bobot": 0.10},
    {"id": 2, "nama": "RO Tidak Lancar",           "rating": 2, "bobot": 0.10},
    {"id": 3, "nama": "RO Lancar",                 "rating": 3, "bobot": 0.10},
    {"id": 4, "nama": "RO Lancar dan Rekomendasi", "rating": 4, "bobot": 0.10},
]

JENIS_JAMINAN = [
    {"id": 1, "nama": "Ijazah Asli Pendidikan Terakhir",       "rating": 4, "bobot": 0.10},
    {"id": 2, "nama": "Ijazah Asli Bukan Pendidikan Terakhir", "rating": 3, "bobot": 0.10},
    {"id": 3, "nama": "Ijazah Legalisir Pendidikan Terakhir",  "rating": 2, "bobot": 0.10},
    {"id": 4, "nama": "Ijazah Legalisir Bukan Pend. Terakhir", "rating": 1, "bobot": 0.10},
]

AKAD = ["Murabahah", "Ijarah", "Musyarakah Mutanaqisah"]

SLIK_KOL = [
    {"kol": 0, "keterangan": "Tidak Ada Fasilitas Pinjaman",  "rating": 4, "bobot": 0.20},
    {"kol": 1, "keterangan": "Lancar",                         "rating": 4, "bobot": 0.20},
    {"kol": 2, "keterangan": "Dalam Pengawasan Khusus (DPK)",  "rating": 3, "bobot": 0.20},
    {"kol": 3, "keterangan": "Kurang Lancar",                  "rating": 2, "bobot": 0.20},
    {"kol": 4, "keterangan": "Diragukan",                      "rating": 1, "bobot": 0.20},
    {"kol": 5, "keterangan": "Macet",                          "rating": 1, "bobot": 0.20},
]

# Score DSR: id → {range_label, rating, bobot} — sesuai database skpd_score_dsrs
DSR_SCORE = [
    {"id": 1, "range": "36% – 40%",     "rating": 1, "bobot": 0.25},
    {"id": 2, "range": "31% – 35%",     "rating": 2, "bobot": 0.25},
    {"id": 3, "range": "21% – 30%",     "rating": 3, "bobot": 0.25},
    {"id": 4, "range": "0% – 20%",      "rating": 4, "bobot": 0.25},
    {"id": 5, "range": "40.01% – 41%",  "rating": 2, "bobot": 0.25},  # range deviasi
    {"id": 6, "range": "< 0% (Negatif)","rating": 1, "bobot": 0.25},
    {"id": 7, "range": "Deviasi DSR",    "rating": 1, "bobot": 0.25},
    {"id": 8, "range": "> 41%",          "rating": 1, "bobot": 0.25},
]

# Kriteria berdasarkan total score terbobot (skala 4, total bobot = 1.00, maks = 4.00)
SCORE_CRITERIA = [
    {"min": 3.25, "max": 4.00, "label": "Sangat Baik",  "status": "DITERIMA"},
    {"min": 2.50, "max": 3.24, "label": "Baik",         "status": "DITERIMA"},
    {"min": 1.75, "max": 2.49, "label": "Cukup",        "status": "DITERIMA"},
    {"min": 1.00, "max": 1.74, "label": "Kurang",       "status": "DITOLAK"},
    {"min": 0.00, "max": 0.99, "label": "Sangat Kurang","status": "DITOLAK"},
]

_NAMA_DEPAN = [
    "Ahmad","Muhammad","Abdul","Hendra","Rudi","Bambang","Agus","Dedi","Eko","Budi",
    "Hadi","Wahyu","Supardi","Tri","Aris","Firmansyah","Joko","Sudirman","Faisal","Asep",
    "Rizki","Hendro","Irwan","Andri","Moch","Siti","Nurul","Fatimah","Dewi","Yuliana",
    "Sri","Rina","Novia","Wulandari","Maya","Lina","Dwi","Endah","Yeni","Sari",
    "Anita","Putri","Nurhayati","Dian","Tuti","Ika","Nita","Susilawati","Retno","Yunita",
    "Rahmad","Syaiful","Hamdani","Ruslan","Surya","Arief","Iwan","Teguh","Yusuf","Zainal",
    "Khairul","Hamid","Darmawan","Saiful","Ridwan","Lukman","Hasanuddin","Basri","Aminuddin","Nor",
    "Halimah","Aisyah","Mardiah","Rukmini","Sumiati","Kartini","Ratna","Fitriani","Indah","Suryani",
    "Nurlaila","Rohana","Aminah","Zulaikha","Hasanah","Ruqayyah","Mariam","Nuraini","Sarinah","Yulianti",
]
_NAMA_BELAKANG = [
    "Fauzi","Rahmah","Karim","Kusuma","Hartono","Supriyadi","Santoso","Kurniawan","Prasetyo","Setiawan",
    "Permana","Nugroho","Wibowo","Fitriani","Munandar","Rahma","Widodo","Hasan","Rahman","Saepudin",
    "Maulana","Effendi","Hidayat","Sholeh","Indriyani","Purnamasari","Handayani","Alawiyah","Purnama",
    "Lestari","Marlina","Anggraini","Susanti","Sari","Apriyanti","Rahayu","Kusumawati","Wulandari",
    "Anwar","Hidayah","Noor","Hamzah","Wahyudi","Siregar","Sitompul","Daulay","Lubis","Hasibuan",
    "Mulia","Pratama","Utama","Saputra","Wijaya","Santosa","Gunawan","Halim","Budiman","Sugiarto",
    "Rasyid","Basuki","Subagyo","Suprayitno","Suharto","Susilo","Prabowo","Sukarno","Winarno","Wardoyo",
    "Yusron","Zainuddin","Mufti","Hakim","Kamaruddin","Salim","Ramli","Taher","Mansur","Arifin",
]

# Generate 300 nama unik
_random_gen = random.Random(99)
NAMA_NASABAH = []
_used = set()
while len(NAMA_NASABAH) < 300:
    dep = _random_gen.choice(_NAMA_DEPAN)
    bel = _random_gen.choice(_NAMA_BELAKANG)
    nm  = f"{dep} {bel}"
    if nm not in _used:
        _used.add(nm)
        NAMA_NASABAH.append(nm)

KOTA = [
    "Banjarmasin", "Banjarbaru", "Martapura", "Rantau", "Kandangan",
    "Barabai", "Amuntai", "Tanjung", "Pelaihari", "Kotabaru",
]

JABATAN_LIST = [
    "Staf Administrasi", "Kepala Seksi", "Kepala Bidang", "Pelaksana",
    "Analis Kebijakan", "Pranata Komputer", "Auditor", "Penyuluh",
    "Tenaga Medis", "Guru Kelas",
]

# ─── Fungsi Pembantu ──────────────────────────────────────────────────────────

def generate_no_ktp(kota_idx, birth_date, gender="L"):
    """Generate KTP-like number (synthetic)"""
    kode = str(random.randint(6300, 6399))
    tgl = birth_date.day
    if gender == "P":
        tgl += 40
    dob_str = f"{tgl:02d}{birth_date.month:02d}{birth_date.year % 100:02d}"
    unik = str(random.randint(1000, 9999))
    return f"{kode}{str(kota_idx).zfill(2)}{dob_str}{unik}"

def random_date_in_2022():
    start = date(2022, 1, 1)
    end   = date(2022, 12, 31)
    delta = end - start
    return start + timedelta(days=random.randint(0, delta.days))

def calculate_dsr(angsuran, golongan_id, gaji_tpp, total_pemasukan):
    """DSR sesuai logika controller SKPD"""
    if golongan_id == 18:  # Non PNS
        return angsuran / total_pemasukan * 100
    else:
        return angsuran / gaji_tpp * 100

def get_dsr_score(dsr, has_deviasi):
    if has_deviasi:
        return DSR_SCORE[6]  # id=7 deviasi → rating 1
    elif 36 <= dsr <= 40:
        return DSR_SCORE[0]  # id=1 → rating 1
    elif 31 <= dsr <= 35:
        return DSR_SCORE[1]  # id=2 → rating 2
    elif 21 <= dsr <= 30:
        return DSR_SCORE[2]  # id=3 → rating 3
    elif 0 <= dsr <= 20:
        return DSR_SCORE[3]  # id=4 → rating 4
    elif dsr < 0:
        return DSR_SCORE[5]  # id=6 → rating 1
    elif 40 < dsr <= 41:
        return DSR_SCORE[4]  # id=5 → rating 2
    else:
        return DSR_SCORE[7]  # id=8 >41% → rating 1

def get_kriteria(total_score):
    for c in SCORE_CRITERIA:
        if c["min"] <= total_score <= c["max"]:
            return c["label"], c["status"]
    return "Tidak Diketahui", "DITOLAK"

def format_rp(val):
    return f"Rp {val:,.0f}".replace(",", ".")

# ─── Generate Dataset ─────────────────────────────────────────────────────────

random.seed(42)

records = []
for i, nama in enumerate(NAMA_NASABAH[:300]):
    # Identitas
    gender = "L" if i % 3 != 1 else "P"
    birth_year  = random.randint(1975, 1998)
    birth_month = random.randint(1, 12)
    birth_day   = random.randint(1, 28)
    tgl_lahir   = date(birth_year, birth_month, birth_day)
    kota_idx    = (i % len(KOTA))
    kota        = KOTA[kota_idx]
    no_ktp      = generate_no_ktp(kota_idx, tgl_lahir, gender)

    # Status & Tanggungan
    status_perkawinan = STATUS_PERKAWINAN[random.choices([0, 1, 2], weights=[15, 75, 10])[0]]
    tanggungan        = TANGGUNGAN[random.choices([0, 1, 2, 3, 4], weights=[20, 30, 25, 15, 10])[0]]

    # Pembiayaan & Instansi
    instansi     = INSTANSI[random.randrange(len(INSTANSI))]
    golongan     = GOLONGAN[random.choices(range(18), weights=[2,2,3,3,5,5,6,6,8,8,9,9,7,7,5,4,3,8])[0]]
    golongan_id  = golongan["id"]
    jabatan      = random.choice(JABATAN_LIST)
    akad         = random.choice(AKAD)

    gaji_pokok        = GAJI_BY_GOLONGAN[golongan_id]
    gaji_tpp          = TPP_BY_GOLONGAN[golongan_id]
    pendapatan_lainnya = random.choice([0, 0, 0, 500000, 750000, 1000000])
    total_pemasukan   = gaji_pokok + gaji_tpp + pendapatan_lainnya

    # Simulasi pengeluaran existing (cicilan SLIK)
    has_existing_slik = random.choices([True, False], weights=[60, 40])[0]
    if has_existing_slik:
        cicilan_slik = random.choice([500000, 750000, 1000000, 1250000, 1500000])
        kol_slik     = random.choices([1, 1, 1, 2, 2, 3, 4, 5], weights=[30, 25, 20, 10, 5, 5, 3, 2])[0]
    else:
        cicilan_slik = 0
        kol_slik     = 0

    pengeluaran_lainnya = random.choice([0, 200000, 300000, 500000])
    biaya_keluarga      = status_perkawinan["biaya"] + tanggungan["biaya"]
    total_pengeluaran   = biaya_keluarga + cicilan_slik + pengeluaran_lainnya
    pendapatan_bersih   = total_pemasukan - total_pengeluaran

    # Pembiayaan yang diajukan
    nominal_ranges = [10000000, 15000000, 20000000, 25000000, 30000000,
                      40000000, 50000000, 75000000, 100000000]
    nominal   = random.choice(nominal_ranges)
    tenor     = random.choice([12, 24, 36, 48, 60, 72, 84])
    rate      = random.choice([9, 10, 11, 12])

    harga_jual = nominal * (rate / 100) * (tenor / 12) + nominal
    angsuran   = harga_jual / tenor
    tgl_pengajuan = random_date_in_2022()

    # DSR
    dsr_val = calculate_dsr(angsuran, golongan_id, gaji_tpp, total_pemasukan)

    # Deviasi DSR (hanya jika DSR 40-41)
    has_deviasi_dsr = (40 < dsr_val <= 41)

    # ── Scoring ──────────────────────────────────────────────────────────────
    # 1. Instansi
    r_instansi   = instansi["rating"]
    b_instansi   = instansi["bobot"]
    n_instansi   = r_instansi * b_instansi

    # 2. Bendahara
    bendahara_obj = random.choices(BENDAHARA, weights=[5, 10, 20, 65])[0]
    r_bendahara   = bendahara_obj["rating"]
    b_bendahara   = bendahara_obj["bobot"]
    n_bendahara   = r_bendahara * b_bendahara

    # 3. DSR
    dsr_score    = get_dsr_score(dsr_val, has_deviasi_dsr)
    r_dsr        = dsr_score["rating"]
    b_dsr        = dsr_score["bobot"]
    n_dsr        = r_dsr * b_dsr

    # 4. SLIK
    slik_entry   = next(s for s in SLIK_KOL if s["kol"] == kol_slik)
    r_slik       = slik_entry["rating"]
    b_slik       = slik_entry["bobot"]
    n_slik       = r_slik * b_slik

    # 5. Jaminan
    jaminan_obj  = random.choices(JENIS_JAMINAN, weights=[40, 30, 20, 10])[0]
    r_jaminan    = jaminan_obj["rating"]
    b_jaminan    = jaminan_obj["bobot"]
    n_jaminan    = r_jaminan * b_jaminan

    # 6. Jenis Nasabah
    jenis_nasabah = random.choices(JENIS_NASABAH, weights=[15, 10, 45, 30])[0]
    r_nasabah    = jenis_nasabah["rating"]
    b_nasabah    = 0.10
    n_nasabah    = r_nasabah * b_nasabah

    # Total Score
    total_score  = n_instansi + n_bendahara + n_dsr + n_slik + n_jaminan + n_nasabah
    kriteria_label, kriteria_status = get_kriteria(total_score)

    records.append({
        "no":                  i + 1,
        "nama_nasabah":        nama,
        "no_ktp":              no_ktp,
        "tempat_lahir":        kota,
        "tgl_lahir":           tgl_lahir.strftime("%d/%m/%Y"),
        "status_perkawinan":   status_perkawinan["nama"],
        "tanggungan":          tanggungan["banyak"],
        "kota":                kota,
        "no_telp":             f"08{random.randint(100000000, 999999999)}",
        "instansi":            instansi["nama"],
        "golongan":            golongan["nama"],
        "jabatan":             jabatan,
        "akad":                akad,
        "tgl_pengajuan":       tgl_pengajuan.strftime("%d/%m/%Y"),
        "nominal_pembiayaan":  nominal,
        "tenor":               tenor,
        "rate":                rate,
        "harga_jual":          round(harga_jual),
        "angsuran":            round(angsuran),
        "gaji_pokok":          gaji_pokok,
        "gaji_tpp":            gaji_tpp,
        "pendapatan_lainnya":  pendapatan_lainnya,
        "total_pemasukan":     total_pemasukan,
        "cicilan_slik":        cicilan_slik,
        "pengeluaran_lainnya": pengeluaran_lainnya,
        "biaya_keluarga":      biaya_keluarga,
        "total_pengeluaran":   total_pengeluaran,
        "pendapatan_bersih":   pendapatan_bersih,
        "dsr":                 round(dsr_val, 2),
        "dsr_range":           dsr_score["range"],
        "has_deviasi_dsr":     "Ya" if has_deviasi_dsr else "Tidak",
        "kol_slik":            kol_slik,
        "keterangan_slik":     slik_entry["keterangan"],
        "jenis_jaminan":       jaminan_obj["nama"],
        "jenis_nasabah":       jenis_nasabah["nama"],
        # Scoring components
        "r_instansi":          r_instansi,  "b_instansi":  b_instansi,  "n_instansi":  round(n_instansi, 4),
        "r_bendahara":         r_bendahara, "b_bendahara": b_bendahara, "n_bendahara": round(n_bendahara, 4),
        "r_dsr":               r_dsr,       "b_dsr":       b_dsr,       "n_dsr":       round(n_dsr, 4),
        "r_slik":              r_slik,      "b_slik":      b_slik,       "n_slik":      round(n_slik, 4),
        "r_jaminan":           r_jaminan,   "b_jaminan":   b_jaminan,   "n_jaminan":   round(n_jaminan, 4),
        "r_nasabah":           r_nasabah,   "b_nasabah":   b_nasabah,   "n_nasabah":   round(n_nasabah, 4),
        "total_score":         round(total_score, 4),
        "kriteria_label":      kriteria_label,
        "kriteria_status":     kriteria_status,
    })

# ─── Build Workbook ───────────────────────────────────────────────────────────

wb = openpyxl.Workbook()

# ── Gaya Umum ──
COLOR_HEADER_DARK   = "1F3864"   # Biru tua
COLOR_HEADER_MED    = "2E75B6"   # Biru medium
COLOR_HEADER_LIGHT  = "BDD7EE"   # Biru muda
COLOR_DITERIMA      = "C6EFCE"   # Hijau muda
COLOR_DITOLAK       = "FFC7CE"   # Merah muda
COLOR_DEVIASI       = "FFEB9C"   # Kuning
COLOR_SUBHEADER     = "D9E1F2"
COLOR_ALT_ROW       = "F2F2F2"
COLOR_WHITE         = "FFFFFF"

FONT_TITLE   = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
FONT_HEADER  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
FONT_DATA    = Font(name="Calibri", size=9)
FONT_BOLD    = Font(name="Calibri", bold=True, size=9)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")

thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

def header_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def style_cell(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font:        cell.font       = font
    if fill:        cell.fill       = fill
    if alignment:   cell.alignment  = alignment
    if border:      cell.border     = border
    if number_format: cell.number_format = number_format

def apply_table_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: DATA NASABAH PEMBIAYAAN
# ═══════════════════════════════════════════════════════════════════════════════

ws1 = wb.active
ws1.title = "Data Nasabah Pembiayaan"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A5"

# Sheet 1 total kolom:
# A(1)=No, B-J(2-10)=Identitas, K-N(11-14)=Pembiayaan, O-V(15-22... wait let me recount
# Cols: 1 No | 2-10 Identitas(9) | 11-14 Pembiayaan(4) | 15-22 Analisa(8) |
#       23-25 SLIK(3) | 26-27 Jaminan+JNasabah(2) |
#       28-33 Scoring Rating(6) | 34=Total | 35=Kriteria | 36=Status
# TOTAL = 36 kolom

# Title
title_cell = ws1.cell(row=1, column=1, value="DATA NASABAH PEMBIAYAAN SKPD TAHUN 2022")
style_cell(title_cell, font=Font(name="Calibri", bold=True, size=14, color="FFFFFF"),
           fill=header_fill(COLOR_HEADER_DARK), alignment=ALIGN_CENTER)
ws1.merge_cells("A1:AJ1")  # 36 kolom
ws1.row_dimensions[1].height = 30

subtitle = ws1.cell(row=2, column=1, value="Bank Pembiayaan Syariah – Data Penelitian")
style_cell(subtitle, font=Font(name="Calibri", italic=True, size=10, color="1F3864"),
           fill=header_fill("EBF3FB"), alignment=ALIGN_CENTER)
ws1.merge_cells("A2:AJ2")
ws1.row_dimensions[2].height = 20

# Kelompok header baris 3
# Cols: 1|2-10|11-14|15-22|23-25|26-27|28-33|34-36
from openpyxl.utils import column_index_from_string
import re as _re

groups = [
    (1,  1,  "No",                      COLOR_HEADER_DARK),
    (2,  10, "IDENTITAS NASABAH",        COLOR_HEADER_DARK),
    (11, 14, "DATA PEMBIAYAAN",          COLOR_HEADER_MED),
    (15, 22, "ANALISA KEUANGAN",         "1F5888"),
    (23, 25, "DATA SLIK",                "1F3864"),
    (26, 27, "JAMINAN & JENIS NASABAH",  "8B4513"),
    (28, 33, "SCORING PER KRITERIA",     "7B2D8B"),
    (34, 36, "KEPUTUSAN",                "1F6B4A"),
]
for (c1, c2, label, color) in groups:
    c = ws1.cell(row=3, column=c1, value=label)
    style_cell(c, font=FONT_HEADER, fill=header_fill(color),
               alignment=ALIGN_CENTER, border=thin_border)
    if c2 > c1:
        ws1.merge_cells(start_row=3, end_row=3, start_column=c1, end_column=c2)
ws1.row_dimensions[3].height = 25

# Sub-header baris 4
sub_headers = [
    # Identitas (1-10)
    "No",
    "Nama Nasabah", "No KTP", "Tempat Lahir", "Tgl Lahir",
    "Status Kawin", "Tanggungan", "Kota/Kab.", "No Telepon",
    "Instansi",
    # Pembiayaan (11-14)
    "Golongan", "Akad", "Nominal (Rp)", "Tgl Pengajuan",
    # Analisa Keuangan (15-22)
    "Tenor (Bln)", "Rate (%/th)", "Harga Jual (Rp)", "Angsuran/Bln (Rp)",
    "Gaji Pokok", "TPP/Tunjangan", "Total Pendapatan", "DSR (%)",
    # SLIK (23-25)
    "KOL SLIK", "Keterangan SLIK", "Deviasi DSR",
    # Jaminan & Jenis Nasabah (26-27)
    "Jenis Jaminan", "Jenis Nasabah",
    # Scoring Per Kriteria (28-33) — rating masing-masing
    "Rating Instansi", "Rating Bendahara", "Rating DSR",
    "Rating SLIK", "Rating Jaminan", "Rating Jenis Nasabah",
    # Keputusan (34-36)
    "Total Score", "Kriteria", "Status",
]

for col_idx, header in enumerate(sub_headers, start=1):
    c = ws1.cell(row=4, column=col_idx, value=header)
    # Warna beda untuk kolom rating
    if 28 <= col_idx <= 33:
        fill_c = header_fill("7B2D8B")
    elif col_idx in [34, 35, 36]:
        fill_c = header_fill("1F6B4A")
    else:
        fill_c = header_fill(COLOR_HEADER_MED)
    style_cell(c, font=FONT_HEADER, fill=fill_c,
               alignment=ALIGN_CENTER, border=thin_border)
ws1.row_dimensions[4].height = 40

# Data rows
# Rating color mapping: 4=hijau, 3=kuning, 2=oranye, 1=merah
_RATING_COLORS = {4:"375623", 3:"7D6608", 2:"974706", 1:"9C0006"}
_RATING_BG     = {4:"C6EFCE", 3:"FFEB9C", 2:"FFDDB3", 1:"FFC7CE"}

for row_idx, rec in enumerate(records, start=5):
    alt = row_idx % 2 == 0
    bg  = header_fill(COLOR_ALT_ROW) if alt else header_fill(COLOR_WHITE)

    row_data = [
        # Identitas (1-10)
        rec["no"], rec["nama_nasabah"], rec["no_ktp"], rec["tempat_lahir"],
        rec["tgl_lahir"], rec["status_perkawinan"], rec["tanggungan"],
        rec["kota"], rec["no_telp"], rec["instansi"],
        # Pembiayaan (11-14)
        rec["golongan"], rec["akad"], rec["nominal_pembiayaan"], rec["tgl_pengajuan"],
        # Analisa Keuangan (15-22)
        rec["tenor"], rec["rate"],
        rec["harga_jual"], rec["angsuran"],
        rec["gaji_pokok"], rec["gaji_tpp"], rec["total_pemasukan"],
        rec["dsr"],
        # SLIK (23-25)
        rec["kol_slik"], rec["keterangan_slik"], rec["has_deviasi_dsr"],
        # Jaminan & Jenis (26-27)
        rec["jenis_jaminan"], rec["jenis_nasabah"],
        # Scoring Per Kriteria (28-33)
        rec["r_instansi"], rec["r_bendahara"], rec["r_dsr"],
        rec["r_slik"],    rec["r_jaminan"],  rec["r_nasabah"],
        # Keputusan (34-36)
        rec["total_score"], rec["kriteria_label"], rec["kriteria_status"],
    ]

    for col_idx, val in enumerate(row_data, start=1):
        c = ws1.cell(row=row_idx, column=col_idx, value=val)
        c.font      = FONT_DATA
        c.fill      = bg
        c.border    = thin_border
        c.alignment = ALIGN_CENTER
        if col_idx in [13, 17, 18, 19, 20, 21]:  # Nominal, harga jual, angsuran, gaji
            c.number_format = '#,##0'
        if col_idx == 22:  # DSR
            c.number_format = '0.00'
        if col_idx == 34:  # Total score
            c.number_format = '0.0000'

    # Warna rating per kriteria (col 28-33)
    for col_idx in range(28, 34):
        rating_val = ws1.cell(row=row_idx, column=col_idx).value
        if rating_val in _RATING_BG:
            ws1.cell(row=row_idx, column=col_idx).fill = header_fill(_RATING_BG[rating_val])
            ws1.cell(row=row_idx, column=col_idx).font = Font(
                name="Calibri", bold=True, size=9, color=_RATING_COLORS[rating_val])

    # Warna status (col 36)
    status_cell = ws1.cell(row=row_idx, column=36)
    if rec["kriteria_status"] == "DITERIMA":
        status_cell.fill = header_fill(COLOR_DITERIMA)
        status_cell.font = Font(name="Calibri", bold=True, size=9, color="375623")
    else:
        status_cell.fill = header_fill(COLOR_DITOLAK)
        status_cell.font = Font(name="Calibri", bold=True, size=9, color="9C0006")

    # Warna kriteria (col 35)
    kriteria_cell = ws1.cell(row=row_idx, column=35)
    if rec["kriteria_status"] == "DITERIMA":
        kriteria_cell.fill = header_fill(COLOR_DITERIMA)
    else:
        kriteria_cell.fill = header_fill(COLOR_DITOLAK)

    # Warna deviasi DSR (col 25)
    if rec["has_deviasi_dsr"] == "Ya":
        ws1.cell(row=row_idx, column=25).fill = header_fill(COLOR_DEVIASI)

    ws1.row_dimensions[row_idx].height = 16

# Column widths sheet 1 (36 kolom)
col_widths_s1 = [
    4,   # 1  No
    22,  # 2  Nama
    18,  # 3  KTP
    13,  # 4  Tempat Lahir
    11,  # 5  Tgl Lahir
    13,  # 6  Status Kawin
    11,  # 7  Tanggungan
    13,  # 8  Kota
    13,  # 9  No Telp
    33,  # 10 Instansi
    14,  # 11 Golongan
    18,  # 12 Akad
    16,  # 13 Nominal
    12,  # 14 Tgl Pengajuan
    9,   # 15 Tenor
    8,   # 16 Rate
    16,  # 17 Harga Jual
    16,  # 18 Angsuran
    14,  # 19 Gaji Pokok
    14,  # 20 TPP
    14,  # 21 Total Pend
    9,   # 22 DSR
    8,   # 23 KOL SLIK
    22,  # 24 Ket SLIK
    10,  # 25 Deviasi DSR
    23,  # 26 Jenis Jaminan
    16,  # 27 Jenis Nasabah
    13,  # 28 Rating Instansi
    14,  # 29 Rating Bendahara
    11,  # 30 Rating DSR
    11,  # 31 Rating SLIK
    13,  # 32 Rating Jaminan
    16,  # 33 Rating JNasabah
    12,  # 34 Total Score
    14,  # 35 Kriteria
    9,   # 36 Status
]
for i, w in enumerate(col_widths_s1, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: SCORING NASABAH
# ═══════════════════════════════════════════════════════════════════════════════

ws2 = wb.create_sheet("Scoring Nasabah")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "A5"

t2 = ws2.cell(row=1, column=1, value="LEMBAR SCORING NASABAH PEMBIAYAAN SKPD TAHUN 2022")
style_cell(t2, font=Font(name="Calibri", bold=True, size=14, color="FFFFFF"),
           fill=header_fill(COLOR_HEADER_DARK), alignment=ALIGN_CENTER)
ws2.merge_cells("A1:V1")
ws2.row_dimensions[1].height = 30

s2 = ws2.cell(row=2, column=1, value="Komponen Scoring: Instansi (10%) | Bendahara (15%) | DSR (25%) | SLIK (20%) | Jaminan (15%) | Jenis Nasabah (10%) = Total 95%")
style_cell(s2, font=Font(name="Calibri", italic=True, size=9, color="1F3864"),
           fill=header_fill("EBF3FB"), alignment=ALIGN_CENTER)
ws2.merge_cells("A2:V2")
ws2.row_dimensions[2].height = 18

# Header kelompok (baris 3)
score_groups = [
    ("A3:A3",  "No"),
    ("B3:B3",  "Nama Nasabah"),
    ("C3:E3",  "INSTANSI (Bobot 10%)"),
    ("F3:H3",  "BENDAHARA (Bobot 15%)"),
    ("I3:K3",  "DSR (Bobot 25%)"),
    ("L3:N3",  "SLIK (Bobot 20%)"),
    ("O3:Q3",  "JAMINAN (Bobot 15%)"),
    ("R3:T3",  "JENIS NASABAH (Bobot 10%)"),
    ("U3:U3",  "Total Score"),
    ("V3:V3",  "Kriteria"),
]
score_colors = [
    COLOR_HEADER_DARK, COLOR_HEADER_DARK,
    "1F6B4A", "1F5888", "7B2D8B", "1F3864",
    "8B4513", "4A5B6B",
    COLOR_HEADER_DARK, COLOR_HEADER_DARK,
]
for (merge_range, label), color in zip(score_groups, score_colors):
    start_ref = merge_range.split(":")[0]
    _m = _re.match(r'([A-Z]+)(\d+)', start_ref)
    _col = column_index_from_string(_m.group(1))
    _row = int(_m.group(2))
    c = ws2.cell(row=_row, column=_col, value=label)
    style_cell(c, font=FONT_HEADER, fill=header_fill(color),
               alignment=ALIGN_CENTER, border=thin_border)
    if merge_range.split(":")[0] != merge_range.split(":")[1]:
        ws2.merge_cells(merge_range)
ws2.row_dimensions[3].height = 25

# Sub-header baris 4
score_subheaders = [
    "No", "Nama Nasabah",
    "Rating", "Bobot", "Nilai",
    "Rating", "Bobot", "Nilai",
    "Rating", "Bobot", "Nilai",
    "Rating", "Bobot", "Nilai",
    "Rating", "Bobot", "Nilai",
    "Rating", "Bobot", "Nilai",
    "Total Score", "Kriteria",
]
for col_idx, sh in enumerate(score_subheaders, start=1):
    c = ws2.cell(row=4, column=col_idx, value=sh)
    style_cell(c, font=FONT_HEADER, fill=header_fill(COLOR_HEADER_MED),
               alignment=ALIGN_CENTER, border=thin_border)
ws2.row_dimensions[4].height = 30

# Data scoring
for row_idx, rec in enumerate(records, start=5):
    alt = row_idx % 2 == 0
    bg  = header_fill(COLOR_ALT_ROW) if alt else header_fill(COLOR_WHITE)

    row_data = [
        rec["no"], rec["nama_nasabah"],
        rec["r_instansi"],  rec["b_instansi"],  rec["n_instansi"],
        rec["r_bendahara"], rec["b_bendahara"], rec["n_bendahara"],
        rec["r_dsr"],       rec["b_dsr"],       rec["n_dsr"],
        rec["r_slik"],      rec["b_slik"],       rec["n_slik"],
        rec["r_jaminan"],   rec["b_jaminan"],   rec["n_jaminan"],
        rec["r_nasabah"],   rec["b_nasabah"],   rec["n_nasabah"],
        rec["total_score"], rec["kriteria_label"],
    ]

    for col_idx, val in enumerate(row_data, start=1):
        c = ws2.cell(row=row_idx, column=col_idx, value=val)
        c.font      = FONT_DATA
        c.fill      = bg
        c.border    = thin_border
        c.alignment = ALIGN_CENTER
        # Format bobot dan nilai sebagai persen/desimal
        if col_idx in [4, 7, 10, 13, 16, 19]:   # bobot
            c.number_format = "0.00"
        if col_idx in [5, 8, 11, 14, 17, 20, 21]: # nilai & total
            c.number_format = "0.0000"

    # Warna rating per kolom rating (col 3,6,9,12,15,18 = offset +1 setiap 3)
    for rating_col in [3, 6, 9, 12, 15, 18]:
        rv = ws2.cell(row=row_idx, column=rating_col).value
        if rv in _RATING_BG:
            ws2.cell(row=row_idx, column=rating_col).fill = header_fill(_RATING_BG[rv])
            ws2.cell(row=row_idx, column=rating_col).font = Font(
                name="Calibri", bold=True, size=9, color=_RATING_COLORS[rv])

    # Warna total score berdasarkan kriteria
    total_c = ws2.cell(row=row_idx, column=21)
    kriteria_c = ws2.cell(row=row_idx, column=22)
    if rec["kriteria_status"] == "DITERIMA":
        if rec["total_score"] >= 4.21:
            col = "375623"; bg_col = "C6EFCE"
        elif rec["total_score"] >= 3.41:
            col = "375623"; bg_col = "D9F0D1"
        else:
            col = "375623"; bg_col = "EBF7E8"
        total_c.fill = header_fill(bg_col)
        kriteria_c.fill = header_fill(bg_col)
        total_c.font = Font(name="Calibri", bold=True, size=9, color=col)
        kriteria_c.font = Font(name="Calibri", bold=True, size=9, color=col)
    else:
        total_c.fill = header_fill(COLOR_DITOLAK)
        kriteria_c.fill = header_fill(COLOR_DITOLAK)
        total_c.font = Font(name="Calibri", bold=True, size=9, color="9C0006")
        kriteria_c.font = Font(name="Calibri", bold=True, size=9, color="9C0006")

    ws2.row_dimensions[row_idx].height = 16

# Lebar kolom sheet 2
col_widths_s2 = [4, 22] + ([9, 7, 8] * 6) + [12, 16]
for i, w in enumerate(col_widths_s2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: TABEL REFERENSI SCORING
# ═══════════════════════════════════════════════════════════════════════════════

ws3 = wb.create_sheet("Referensi Scoring")
ws3.sheet_view.showGridLines = False

def write_ref_table(ws, start_row, start_col, title, headers, rows, title_color="1F3864", header_color="2E75B6"):
    # Title
    end_col = start_col + len(headers) - 1
    merge = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{start_row}"
    c = ws.cell(row=start_row, column=start_col, value=title)
    style_cell(c, font=Font(name="Calibri", bold=True, size=11, color="FFFFFF"),
               fill=header_fill(title_color), alignment=ALIGN_CENTER, border=thin_border)
    if end_col > start_col:
        ws.merge_cells(merge)
    ws.row_dimensions[start_row].height = 22

    # Header baris
    for j, h in enumerate(headers):
        c = ws.cell(row=start_row+1, column=start_col+j, value=h)
        style_cell(c, font=FONT_HEADER, fill=header_fill(header_color),
                   alignment=ALIGN_CENTER, border=thin_border)
    ws.row_dimensions[start_row+1].height = 30

    # Data baris
    for ri, row_data in enumerate(rows):
        alt_bg = COLOR_ALT_ROW if ri % 2 == 0 else COLOR_WHITE
        for j, val in enumerate(row_data):
            c = ws.cell(row=start_row+2+ri, column=start_col+j, value=val)
            style_cell(c, font=FONT_DATA, fill=header_fill(alt_bg),
                       alignment=ALIGN_CENTER, border=thin_border)
        ws.row_dimensions[start_row+2+ri].height = 18

    return start_row + 2 + len(rows) + 2  # next available row

t3 = ws3.cell(row=1, column=1, value="TABEL REFERENSI SISTEM SCORING PEMBIAYAAN SKPD 2022")
style_cell(t3, font=Font(name="Calibri", bold=True, size=13, color="FFFFFF"),
           fill=header_fill(COLOR_HEADER_DARK), alignment=ALIGN_CENTER)
ws3.merge_cells("A1:N1")
ws3.row_dimensions[1].height = 28

# ── Tabel Scoring DSR ──
next_row = write_ref_table(
    ws3, 3, 1,
    "SCORING DSR (Debt Service Ratio) – Bobot 25%",
    ["ID", "Rentang DSR", "Rating", "Bobot", "Nilai Maks", "Keterangan"],
    [
        [1, "36% – 40%",      1, 0.25, 1*0.25, "Kurang"],
        [2, "31% – 35%",      2, 0.25, 2*0.25, "Cukup"],
        [3, "21% – 30%",      3, 0.25, 3*0.25, "Baik"],
        [4, "0% – 20%",       4, 0.25, 4*0.25, "Sangat Baik"],
        [5, "40.01% – 41%",   2, 0.25, 2*0.25, "Range Pengajuan Deviasi"],
        [6, "< 0% (Negatif)", 1, 0.25, 1*0.25, "Tidak Layak"],
        [7, "Deviasi DSR",    1, 0.25, 1*0.25, "Dengan Persetujuan Deviasi"],
        [8, "> 41%",          1, 0.25, 1*0.25, "Tidak Layak / Tolak"],
    ],
    title_color="7B2D8B"
)

# ── Tabel Scoring SLIK ──
next_row = write_ref_table(
    ws3, next_row, 1,
    "SCORING SLIK / iDEB (Kualitas Kredit) – Bobot 20%",
    ["KOL", "Keterangan", "Rating", "Bobot", "Nilai Maks", "Keterangan Umum"],
    [
        [0, "Tidak Ada Fasilitas",           4, 0.20, 4*0.20, "Sangat Baik"],
        [1, "Lancar",                          4, 0.20, 4*0.20, "Sangat Baik"],
        [2, "Dalam Pengawasan Khusus (DPK)",  3, 0.20, 3*0.20, "Baik"],
        [3, "Kurang Lancar",                  2, 0.20, 2*0.20, "Cukup / Perlu Perhatian"],
        [4, "Diragukan",                      1, 0.20, 1*0.20, "Kurang / Risiko Tinggi"],
        [5, "Macet",                          1, 0.20, 1*0.20, "Tidak Layak"],
    ],
    title_color="1F3864"
)

# ── Tabel Scoring Instansi ──
next_row = write_ref_table(
    ws3, next_row, 1,
    "SCORING INSTANSI – Bobot 10%",
    ["Nama Instansi", "Rating", "Bobot", "Nilai Maks"],
    [[i["nama"], i["rating"], i["bobot"], i["rating"]*i["bobot"]] for i in INSTANSI],
    title_color="1F6B4A"
)

# ── Tabel Scoring Bendahara ──
next_row = write_ref_table(
    ws3, next_row, 1,
    "SCORING BENDAHARA – Bobot 30%",
    ["Rating", "Bobot", "Nilai Maks", "Keterangan Bendahara"],
    [
        [4, 0.30, 4*0.30, "Baik dan Tertib"],
        [3, 0.30, 3*0.30, "Tidak Baik dan Tertib"],
        [2, 0.30, 2*0.30, "Baik dan Tidak Tertib"],
        [1, 0.30, 1*0.30, "Tidak Baik dan Tidak Tertib"],
    ],
    title_color="1F5888"
)

# ── Tabel Scoring Jaminan ──
next_row = write_ref_table(
    ws3, next_row, 1,
    "SCORING JENIS JAMINAN – Bobot 10%",
    ["Jenis Jaminan", "Rating", "Bobot", "Nilai Maks"],
    [[j["nama"], j["rating"], j["bobot"], j["rating"]*j["bobot"]] for j in JENIS_JAMINAN],
    title_color="8B4513"
)

# ── Tabel Scoring Jenis Nasabah ──
next_row = write_ref_table(
    ws3, next_row, 1,
    "SCORING JENIS NASABAH – Bobot 10%",
    ["Jenis Nasabah", "Rating", "Bobot", "Nilai Maks"],
    [[n["nama"], n["rating"], n["bobot"], n["rating"]*n["bobot"]] for n in JENIS_NASABAH],
    title_color="4A5B6B"
)

# ── Tabel Kriteria Keputusan ──
next_row = write_ref_table(
    ws3, next_row, 1,
    "KRITERIA KEPUTUSAN BERDASARKAN TOTAL SCORE TERBOBOT (Skala 4, Total Bobot = 1.00, Maks = 4.00)",
    ["Score Min", "Score Max", "Predikat", "Status", "Keterangan"],
    [
        [3.25, 4.00, "Sangat Baik",  "DITERIMA", "Disetujui tanpa syarat"],
        [2.50, 3.24, "Baik",         "DITERIMA", "Disetujui"],
        [1.75, 2.49, "Cukup",        "DITERIMA", "Disetujui dengan syarat/pertimbangan"],
        [1.00, 1.74, "Kurang",       "DITOLAK",  "Perlu analisa ulang atau tolak"],
        [0.00, 0.99, "Sangat Kurang","DITOLAK",  "Ditolak"],
    ],
    title_color=COLOR_HEADER_DARK
)

# Warna pada tabel kriteria
for r in range(next_row - 5, next_row):
    status_c = ws3.cell(row=r, column=4)
    if status_c.value == "DITERIMA":
        status_c.fill = header_fill(COLOR_DITERIMA)
        status_c.font = Font(name="Calibri", bold=True, size=9, color="375623")
    elif status_c.value == "DITOLAK":
        status_c.fill = header_fill(COLOR_DITOLAK)
        status_c.font = Font(name="Calibri", bold=True, size=9, color="9C0006")

# Lebar kolom sheet 3
for i, w in enumerate([8, 35, 8, 8, 10, 30], start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4: REKAPITULASI
# ═══════════════════════════════════════════════════════════════════════════════

ws4 = wb.create_sheet("Rekapitulasi")
ws4.sheet_view.showGridLines = False

t4 = ws4.cell(row=1, column=1, value="REKAPITULASI DATA NASABAH PEMBIAYAAN SKPD 2022")
style_cell(t4, font=Font(name="Calibri", bold=True, size=13, color="FFFFFF"),
           fill=header_fill(COLOR_HEADER_DARK), alignment=ALIGN_CENTER)
ws4.merge_cells("A1:H1")
ws4.row_dimensions[1].height = 28

# Hitung statistik
total_all    = len(records)
total_terima = sum(1 for r in records if r["kriteria_status"] == "DITERIMA")
total_tolak  = total_all - total_terima
total_nominal = sum(r["nominal_pembiayaan"] for r in records if r["kriteria_status"] == "DITERIMA")
avg_score = sum(r["total_score"] for r in records) / total_all
avg_dsr   = sum(r["dsr"] for r in records) / total_all

# Distribusi status
sangat_baik  = sum(1 for r in records if r["kriteria_label"] == "Sangat Baik")
baik         = sum(1 for r in records if r["kriteria_label"] == "Baik")
cukup        = sum(1 for r in records if r["kriteria_label"] == "Cukup")
kurang       = sum(1 for r in records if r["kriteria_label"] == "Kurang")
sangat_kurang = sum(1 for r in records if r["kriteria_label"] == "Sangat Kurang")

summary_data = [
    ["RINGKASAN KEPUTUSAN PEMBIAYAAN", ""],
    ["Total Pengajuan",                total_all],
    ["Total Diterima",                 total_terima],
    ["Total Ditolak",                  total_tolak],
    ["Rasio Persetujuan",               f"{total_terima/total_all*100:.1f}%"],
    ["Total Nominal Disetujui",        total_nominal],
    ["Rata-rata Score",                round(avg_score, 4)],
    ["Rata-rata DSR",                  f"{avg_dsr:.2f}%"],
    ["", ""],
    ["DISTRIBUSI PREDIKAT SCORING", ""],
    ["Sangat Baik (4.21 – 5.00)",      sangat_baik],
    ["Baik (3.41 – 4.20)",             baik],
    ["Cukup (2.61 – 3.40)",            cukup],
    ["Kurang (1.81 – 2.60)",           kurang],
    ["Sangat Kurang (1.00 – 1.80)",    sangat_kurang],
    ["", ""],
    ["DISTRIBUSI GOLONGAN NASABAH", ""],
]

# Distribusi golongan
gol_counter = {}
for r in records:
    g = r["golongan"]
    gol_counter[g] = gol_counter.get(g, 0) + 1
for gol_name, cnt in sorted(gol_counter.items()):
    summary_data.append([gol_name, cnt])

summary_data += [
    ["", ""],
    ["DISTRIBUSI INSTANSI PENGAJU", ""],
]
inst_counter = {}
for r in records:
    inst = r["instansi"]
    inst_counter[inst] = inst_counter.get(inst, 0) + 1
for inst_name, cnt in sorted(inst_counter.items()):
    summary_data.append([inst_name, cnt])

for row_idx, (label, val) in enumerate(summary_data, start=3):
    c_label = ws4.cell(row=row_idx, column=1, value=label)
    c_val   = ws4.cell(row=row_idx, column=2, value=val)

    if label in ("RINGKASAN KEPUTUSAN PEMBIAYAAN", "DISTRIBUSI PREDIKAT SCORING",
                 "DISTRIBUSI GOLONGAN NASABAH", "DISTRIBUSI INSTANSI PENGAJU"):
        style_cell(c_label, font=Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
                   fill=header_fill(COLOR_HEADER_MED), alignment=ALIGN_CENTER, border=thin_border)
        ws4.merge_cells(f"A{row_idx}:H{row_idx}")
        ws4.row_dimensions[row_idx].height = 22
    elif label == "":
        ws4.row_dimensions[row_idx].height = 8
    else:
        c_label.font   = FONT_BOLD
        c_label.border = thin_border
        c_label.alignment = ALIGN_LEFT
        c_val.font     = FONT_DATA
        c_val.border   = thin_border
        c_val.alignment = ALIGN_CENTER
        ws4.row_dimensions[row_idx].height = 18

        if label in ("Total Diterima",):
            c_val.fill = header_fill(COLOR_DITERIMA)
            c_val.font = Font(name="Calibri", bold=True, size=9, color="375623")
        elif label in ("Total Ditolak",):
            c_val.fill = header_fill(COLOR_DITOLAK)
            c_val.font = Font(name="Calibri", bold=True, size=9, color="9C0006")
        elif label == "Total Nominal Disetujui":
            c_val.number_format = '#,##0'

ws4.column_dimensions["A"].width = 38
ws4.column_dimensions["B"].width = 20

# ─── Simpan File ──────────────────────────────────────────────────────────────

output_path = "/Users/ajspryn/Project/fos/storage/app/Data_Nasabah_SKPD_2022.xlsx"
wb.save(output_path)
print(f"✓ File berhasil disimpan: {output_path}")
print(f"  Total nasabah  : {total_all}")
print(f"  Total diterima : {total_terima}")
print(f"  Total ditolak  : {total_tolak}")
print(f"  Nominal disetujui: Rp {total_nominal:,.0f}")
